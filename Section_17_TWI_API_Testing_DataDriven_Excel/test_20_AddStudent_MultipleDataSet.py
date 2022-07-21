# /**
# * @author Rajat Verma
# * https://www.linkedin.com/in/rajat-v-3b0685128/
# * https://github.com/rajatt95
# * https://rajatt95.github.io/
# * https://github.com/stars/rajatt95/lists/udemy-testing-world-infotech
# * https://github.com/rajatt95/Python_Requests_TWI
# *
# * Course: Step by Step Rest API Testing using Python + Pytest +Allure (https://www.udemy.com/course/api-testing-python/)
# * Tutor: Testing World Infotech (https://www.udemy.com/user/technology-world/)
# */
#
# /***************************************************/
import openpyxl
import requests
import json

# URL
base_url = "https://thetestingworldapi.com/"


def test_01_POST_DataDriven_MultipleDataSets_Using_XLSX_File():
    # URL
    url = base_url + "api/studentsDetails"

    # Open the JSON file in READ mode
    file = open('../files/03_DataDriven_Excel/01_AddStudent.json', 'r')

    # Read the content from the file
    input_JSON = file.read()

    # Convert txt format to JSON format
    requestBody_JSON = json.loads(input_JSON)

    # TestData.xlsx -> Name of the XLSX file
    workbook = openpyxl.load_workbook('../files/03_DataDriven_Excel/TestData.xlsx')

    # AddStudent -> Sheet name
    sheet_addStudent = workbook['AddStudent']

    # Rows count
    rows = sheet_addStudent.max_row

    # Not starting from Row 1 (because 1st row is Header)
    for index in range(2, rows + 1):
        # Column #1 has First name in XLSX sheet
        cell_firstName = sheet_addStudent.cell(row=index, column=1)
        cell_middleName = sheet_addStudent.cell(row=index, column=2)
        cell_lastName = sheet_addStudent.cell(row=index, column=3)
        cell_dateOfBirth = sheet_addStudent.cell(row=index, column=4)

        # Picked the data from XLSX sheet ->
        # Now, putting that Data into JSON file
        requestBody_JSON['first_name'] = cell_firstName.value
        requestBody_JSON['middle_name'] = cell_middleName.value
        requestBody_JSON['last_name'] = cell_lastName.value
        requestBody_JSON['date_of_birth'] = cell_dateOfBirth.value

        print(requestBody_JSON)

        # Send POST Request with Request body
        response = requests.post(url, requestBody_JSON)

        print(
            "===========>>>>>>>>>>>>> response.text ==>> " + response.text)

        # Assertion for Response Status Code
        assert response.status_code == 201
