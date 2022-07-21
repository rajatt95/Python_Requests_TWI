
import openpyxl


class CommonUtilities:

    # Constructor
    def __init__(self, fileNamePath, sheetName):
        # Making these 2 variables as GLOBAL
        # Because they will be used in many methods
        global workbook
        global sheet

        workbook = openpyxl.load_workbook(fileNamePath)
        sheet = workbook[sheetName]

    def fetch_row_count(self):
        row_count = sheet.max_row  # Rows count
        return row_count

    def fetch_column_count(self):
        column_count = sheet.max_column # Columns count
        return column_count

    def fetch_key_names(self):
        column_count = sheet.max_column   # Columns count
        test_data_list = []  # Empty List
        for index in range(1, column_count+1):
            cell = sheet.cell(row=1, column=index)
            test_data_list.insert(index-1, cell.value)
        return test_data_list

    def update_request_with_data(self,row_number, json_request, key_list):
        column_count = sheet.max_column # Columns count
        for index in range(1, column_count+1):
            cell = sheet.cell(row=row_number, column=index)
            json_request[key_list[index-1]] = cell.value
        return json_request



